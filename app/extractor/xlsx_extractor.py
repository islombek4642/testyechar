"""
XLSX (Excel) text extractor.

Reads Excel spreadsheets containing test questions in two main layouts, 
with smart automatic detection of metadata/subject columns (e.g. "Fan nomi" or Question Numbers in Column A),
and dynamic header detection supporting merged title cells in the first row.

Converts them into a clean CLASSIC format text structure for parsing.
"""

from __future__ import annotations

import openpyxl
from pathlib import Path
from app.utils.logger import get_logger
from .base import BaseExtractor, ExtractionResult

log = get_logger(__name__)


class XLSXExtractor(BaseExtractor):
    """
    Extractor for Excel (.xlsx) files.
    
    Converts Excel test configurations into classic format text (? / + / =).
    """

    @property
    def name(self) -> str:
        return "XLSXExtractor"

    def supports(self, path: Path) -> bool:
        """Return True if the file has a .xlsx extension."""
        return path.suffix.lower() == ".xlsx"

    def extract(self, path: Path) -> ExtractionResult:
        """
        Extract questions from Excel sheets and build classic format text.
        
        Args:
            path: Path to the XLSX file.
            
        Returns:
            ExtractionResult containing the formatted text.
        """
        log.info(f"Excel faylidan ma'lumotlar o'qilmoqda: {path.name}")
        
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True)
        except Exception as e:
            log.error(f"Excel faylini ochishda xatolik yuz berdi: {e}", exc_info=True)
            return ExtractionResult(
                pages=[],
                page_count=0,
                source_path=path,
                extractor_name=self.name
            )

        questions = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            log.info(f"Varaq tahlil qilinmoqda: {sheet_name}")
            
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
                
            # --- DETECT THE FIRST ROW THAT CONTAINS ACTUAL MULTI-COLUMN DATA (>= 3 columns) ---
            # This is essential for sheets where row 1 is a merged title cell and row 2 contains the headers.
            first_valid_row_idx = None
            for row_idx, row in enumerate(rows, start=1):
                row_values = list(row)
                while row_values and row_values[-1] is None:
                    row_values.pop()
                if not row_values:
                    continue
                non_empty = [val for val in row_values if val is not None and str(val).strip()]
                if len(non_empty) >= 3:
                    first_valid_row_idx = row_idx
                    break
                    
            log.info(f"Birinchi yaroqli ma'lumotlar qatori indeksi: {first_valid_row_idx}")
            
            # --- PRE-ANALYSIS OF THE SHEET TO DETECT COLUMN A CHARACTERISTICS ---
            # We want to identify if Column A represents metadata (e.g. Subject Name, Question Number)
            # or if it contains the actual Question Text.
            valid_rows_data = []
            
            for row_idx, row in enumerate(rows, start=1):
                row_values = list(row)
                
                # Trim trailing None values
                while row_values and row_values[-1] is None:
                    row_values.pop()
                    
                if not row_values:
                    continue
                    
                # Get non-empty cells
                non_empty = [(i, val) for i, val in enumerate(row_values) if val is not None and str(val).strip()]
                if len(non_empty) < 3:
                    # Rows with < 3 non-empty cells (like a merged title in Row 1) are skipped
                    continue
                    
                # Header check - Only check the very first valid data row containing >= 3 columns
                is_header = False
                if row_idx == first_valid_row_idx:
                    first_vals = [str(x[1]).lower() for x in non_empty[:3]]
                    for v in first_vals:
                        words = v.replace("/", " ").replace(".", " ").replace("-", " ").split()
                        if any(w in ["savol", "question", "to'g'ri", "noto'g'ri", "correct", "incorrect", "t/r", "№", "id", "javob", "fan"] for w in words):
                            is_header = True
                            break
                        if any(phrase in v for phrase in ["savol matni", "to'g'ri javob", "noto'g'ri javob", "savol_matni", "savol matn", "fan nomi", "fan_nomi"]):
                            is_header = True
                            break
                if is_header:
                    log.info(f"Sarlavha qatori aniqlandi va o'tkazib yuborildi (satr {row_idx}): {row_values[:3]}")
                    continue
                    
                valid_rows_data.append((row_idx, row_values))
                
            if not valid_rows_data:
                continue
                
            # Gather all non-empty Column A values to analyze
            col_a_vals = []
            for _, r_vals in valid_rows_data:
                if r_vals[0] is not None:
                    col_a_vals.append(str(r_vals[0]).strip())
            
            # Default to False
            is_col_a_metadata = False
            
            if col_a_vals:
                unique_a = set(col_a_vals)
                ratio_unique = len(unique_a) / len(col_a_vals)
                
                # Case 1: Column A contains a repeating subject name or category (high duplication rate)
                if len(unique_a) == 1 or ratio_unique < 0.30:
                    is_col_a_metadata = True
                    log.info(f"A ustuni takrorlanuvchi metadata (masalan, Fan nomi) deb aniqlandi. Unique ratio: {ratio_unique:.2f}")
                else:
                    # Case 2: Column A contains question numbers/IDs (all values are short, numeric, or contain keywords like 'Savol')
                    all_short_or_numeric = True
                    for val in col_a_vals:
                        is_num = val.isdigit()
                        is_short = len(val) <= 10
                        is_q_num = False
                        
                        if not is_num:
                            clean_val = val.replace(".", "").replace("-", "").strip()
                            if clean_val.isdigit():
                                is_q_num = True
                            elif any(x in val.lower() for x in ["savol", "quest", "t/r", "№"]):
                                is_q_num = True
                                
                        if not (is_num or is_short or is_q_num):
                            all_short_or_numeric = False
                            break
                            
                    if all_short_or_numeric:
                        is_col_a_metadata = True
                        log.info("A ustuni savol raqamlari/indekslari deb aniqlandi.")
            
            # --- PROCESS THE ROWS ACCORDING TO THE DETECTED LAYOUT ---
            for row_idx, row_values in valid_rows_data:
                if is_col_a_metadata:
                    # Layout 1 (A = Metadata/Subject/No, B = Question, C = Correct, D+ = Incorrect)
                    q_text = str(row_values[1]).strip() if len(row_values) > 1 and row_values[1] is not None else ""
                    correct_ans = str(row_values[2]).strip() if len(row_values) > 2 and row_values[2] is not None else ""
                    wrongs = []
                    for val in row_values[3:]:
                        if val is not None:
                            val_str = str(val).strip()
                            if val_str:
                                wrongs.append(val_str)
                else:
                    # Layout 2 (A = Question, B = Correct, C+ = Incorrect)
                    q_text = str(row_values[0]).strip() if row_values[0] is not None else ""
                    correct_ans = str(row_values[1]).strip() if len(row_values) > 1 and row_values[1] is not None else ""
                    wrongs = []
                    for val in row_values[2:]:
                        if val is not None:
                            val_str = str(val).strip()
                            if val_str:
                                wrongs.append(val_str)
                                
                if q_text and correct_ans and wrongs:
                    questions.append({
                        "question": q_text,
                        "correct": correct_ans,
                        "wrongs": wrongs,
                        "row_idx": row_idx,
                        "sheet_name": sheet_name
                    })

        # Build Classic format text from the extracted questions list
        classic_lines = []
        for q in questions:
            classic_lines.append(f"? {q['question']}")
            classic_lines.append(f"+ {q['correct']}")
            for w in q['wrongs']:
                classic_lines.append(f"= {w}")
            classic_lines.append("") # Empty line separator
            
        full_text = "\n".join(classic_lines)
        log.info(f"Excel faylidan {len(questions)} ta yaroqli savol ajratib olindi.")
        
        # We wrap the full classic text inside one single "page"
        return ExtractionResult(
            pages=[full_text] if full_text else [],
            page_count=1 if full_text else 0,
            source_path=path,
            extractor_name=self.name
        )
